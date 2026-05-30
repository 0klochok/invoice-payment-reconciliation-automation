"""Input ingestion, normalization, and validation for invoice/payment files."""

from __future__ import annotations

import csv
import re
from collections.abc import Iterator
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from .models import (
    ImportDiagnostics,
    ImportResult,
    InvoiceRecord,
    MoneyAmount,
    PaymentRecord,
    ValidationError,
)

INVOICE_REQUIRED_FIELDS = (
    "invoice_id",
    "customer_name",
    "invoice_date",
    "due_date",
    "invoice_amount",
    "currency",
)
PAYMENT_REQUIRED_FIELDS = (
    "payment_id",
    "customer_name",
    "payment_date",
    "payment_amount",
    "currency",
    "payment_reference",
)

MISSING_REQUIRED = "missing_required"
INVALID_DATE = "invalid_date"
INVALID_AMOUNT = "invalid_amount"

MISSING_REQUIRED_MESSAGE = "Missing required value."
INVALID_DATE_MESSAGE = "Expected ISO date in YYYY-MM-DD format."
INVALID_AMOUNT_MESSAGE = "Expected decimal money amount."
UNSUPPORTED_INPUT_MESSAGE = "Expected a .csv or .xlsx input file."

_ISO_DATE_PATTERN = re.compile(r"\d{4}-\d{2}-\d{2}\Z")


def load_invoice_file(path: str | Path) -> ImportResult[InvoiceRecord]:
    """Load, normalize, and validate invoice records from CSV or XLSX input."""
    file_path = Path(path)
    if file_path.suffix.lower() == ".csv":
        return load_invoice_csv(file_path)
    if file_path.suffix.lower() == ".xlsx":
        return load_invoice_xlsx(file_path)
    raise ValueError(UNSUPPORTED_INPUT_MESSAGE)


def load_payment_file(path: str | Path) -> ImportResult[PaymentRecord]:
    """Load, normalize, and validate payment records from CSV or XLSX input."""
    file_path = Path(path)
    if file_path.suffix.lower() == ".csv":
        return load_payment_csv(file_path)
    if file_path.suffix.lower() == ".xlsx":
        return load_payment_xlsx(file_path)
    raise ValueError(UNSUPPORTED_INPUT_MESSAGE)


def load_invoice_csv(path: str | Path) -> ImportResult[InvoiceRecord]:
    """Load, normalize, and validate invoice records from a CSV file."""
    return _load_invoices(_read_csv(path))


def load_payment_csv(path: str | Path) -> ImportResult[PaymentRecord]:
    """Load, normalize, and validate payment records from a CSV file."""
    return _load_payments(_read_csv(path))


def load_invoice_xlsx(path: str | Path) -> ImportResult[InvoiceRecord]:
    """Load, normalize, and validate invoice records from an XLSX file."""
    return _load_invoices(_read_xlsx(path))


def load_payment_xlsx(path: str | Path) -> ImportResult[PaymentRecord]:
    """Load, normalize, and validate payment records from an XLSX file."""
    return _load_payments(_read_xlsx(path))


def _load_invoices(
    rows: Iterator[tuple[int, dict[str, str | None]]],
) -> ImportResult[InvoiceRecord]:
    records: list[InvoiceRecord] = []
    errors: list[ValidationError] = []
    rows_seen = 0

    for row_number, row in rows:
        rows_seen += 1
        record, row_errors = _parse_invoice_row(row, row_number)
        errors.extend(row_errors)
        if record is not None:
            records.append(record)

    return ImportResult(
        records=tuple(records),
        diagnostics=_build_diagnostics("invoices", rows_seen, len(records), errors),
    )


def _load_payments(
    rows: Iterator[tuple[int, dict[str, str | None]]],
) -> ImportResult[PaymentRecord]:
    records: list[PaymentRecord] = []
    errors: list[ValidationError] = []
    rows_seen = 0

    for row_number, row in rows:
        rows_seen += 1
        record, row_errors = _parse_payment_row(row, row_number)
        errors.extend(row_errors)
        if record is not None:
            records.append(record)

    return ImportResult(
        records=tuple(records),
        diagnostics=_build_diagnostics("payments", rows_seen, len(records), errors),
    )


def _read_csv(path: str | Path) -> Iterator[tuple[int, dict[str, str | None]]]:
    with Path(path).open("r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.DictReader(csv_file)
        yield from enumerate(reader, start=2)


def _read_xlsx(path: str | Path) -> Iterator[tuple[int, dict[str, str | None]]]:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = tuple(_cell_to_text(value) or "" for value in next(rows))
        except StopIteration:
            return

        for row_number, values in enumerate(rows, start=2):
            yield (
                row_number,
                {
                    header: _cell_to_text(value)
                    for header, value in zip(headers, values, strict=False)
                    if header
                },
            )
    finally:
        workbook.close()


def _cell_to_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _parse_invoice_row(
    row: dict[str, str | None],
    row_number: int,
) -> tuple[InvoiceRecord | None, list[ValidationError]]:
    source = "invoices"
    values = _normalize_values(row, INVOICE_REQUIRED_FIELDS)
    missing_fields, errors = _required_errors(source, row_number, values)

    invoice_date = _validated_date(
        source, row_number, "invoice_date", values["invoice_date"], missing_fields
    )
    due_date = _validated_date(
        source, row_number, "due_date", values["due_date"], missing_fields
    )
    invoice_amount = _validated_money(
        source,
        row_number,
        "invoice_amount",
        values["invoice_amount"],
        values["currency"],
        missing_fields,
    )

    for _, error in (invoice_date, due_date, invoice_amount):
        if error is not None:
            errors.append(error)
    if errors:
        return None, errors

    return (
        InvoiceRecord(
            invoice_id=values["invoice_id"],
            customer_name=values["customer_name"],
            invoice_date=invoice_date[0],
            due_date=due_date[0],
            invoice_amount=invoice_amount[0],
            source_row=row_number,
        ),
        errors,
    )


def _parse_payment_row(
    row: dict[str, str | None],
    row_number: int,
) -> tuple[PaymentRecord | None, list[ValidationError]]:
    source = "payments"
    values = _normalize_values(row, PAYMENT_REQUIRED_FIELDS)
    missing_fields, errors = _required_errors(source, row_number, values)

    payment_date = _validated_date(
        source, row_number, "payment_date", values["payment_date"], missing_fields
    )
    payment_amount = _validated_money(
        source,
        row_number,
        "payment_amount",
        values["payment_amount"],
        values["currency"],
        missing_fields,
    )

    for _, error in (payment_date, payment_amount):
        if error is not None:
            errors.append(error)
    if errors:
        return None, errors

    return (
        PaymentRecord(
            payment_id=values["payment_id"],
            customer_name=values["customer_name"],
            payment_date=payment_date[0],
            payment_amount=payment_amount[0],
            payment_reference=values["payment_reference"],
            source_row=row_number,
        ),
        errors,
    )


def _normalize_values(
    row: dict[str, str | None],
    required_fields: tuple[str, ...],
) -> dict[str, str]:
    values = {field: _clean_text(row.get(field)) for field in required_fields}
    if "currency" in values:
        values["currency"] = values["currency"].upper()
    return values


def _clean_text(value: str | None) -> str:
    return "" if value is None else value.strip()


def _required_errors(
    source: str,
    row_number: int,
    values: dict[str, str],
) -> tuple[set[str], list[ValidationError]]:
    missing_fields = {field for field, value in values.items() if value == ""}
    errors = [
        _field_error(
            source, row_number, field, MISSING_REQUIRED, MISSING_REQUIRED_MESSAGE
        )
        for field in values
        if field in missing_fields
    ]
    return missing_fields, errors


def _validated_date(
    source: str,
    row_number: int,
    field: str,
    value: str,
    missing_fields: set[str],
) -> tuple[date | None, ValidationError | None]:
    if field in missing_fields:
        return None, None
    if not _ISO_DATE_PATTERN.fullmatch(value):
        return None, _field_error(
            source, row_number, field, INVALID_DATE, INVALID_DATE_MESSAGE
        )
    try:
        return date.fromisoformat(value), None
    except ValueError:
        return None, _field_error(
            source, row_number, field, INVALID_DATE, INVALID_DATE_MESSAGE
        )


def _validated_money(
    source: str,
    row_number: int,
    field: str,
    value: str,
    currency: str,
    missing_fields: set[str],
) -> tuple[MoneyAmount | None, ValidationError | None]:
    if field in missing_fields:
        return None, None
    try:
        parsed = Decimal(value)
    except InvalidOperation:
        return None, _field_error(
            source, row_number, field, INVALID_AMOUNT, INVALID_AMOUNT_MESSAGE
        )
    if not parsed.is_finite():
        return None, _field_error(
            source, row_number, field, INVALID_AMOUNT, INVALID_AMOUNT_MESSAGE
        )
    return MoneyAmount(value=parsed, currency=currency), None


def _field_error(
    source: str,
    row_number: int,
    field: str,
    error_code: str,
    message: str,
) -> ValidationError:
    return ValidationError(
        source=source,
        row_number=row_number,
        field=field,
        error_code=error_code,
        message=message,
    )


def _build_diagnostics(
    source: str,
    rows_seen: int,
    valid_rows: int,
    errors: list[ValidationError],
) -> ImportDiagnostics:
    invalid_rows = len({error.row_number for error in errors})
    return ImportDiagnostics(
        source=source,
        rows_seen=rows_seen,
        valid_rows=valid_rows,
        invalid_rows=invalid_rows,
        errors=tuple(errors),
    )
