"""Local deterministic reconciliation report generation."""

from __future__ import annotations

import csv
from collections.abc import Iterable
from dataclasses import dataclass
from decimal import Decimal
from io import StringIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .matching import (
    AmbiguousReference,
    AmbiguousReferenceReason,
    AmountMismatch,
    CurrencyMismatch,
    MatchedPair,
    MatchStatus,
    ReconciliationResult,
    UnmatchedInvoice,
    UnmatchedPayment,
)
from .models import InvoiceRecord, MoneyAmount, PaymentRecord

STATUS_ORDER = (
    MatchStatus.MATCHED,
    MatchStatus.UNMATCHED_INVOICE,
    MatchStatus.UNMATCHED_PAYMENT,
    MatchStatus.AMOUNT_MISMATCH,
    MatchStatus.CURRENCY_MISMATCH,
    MatchStatus.AMBIGUOUS_REFERENCE,
)

STATUS_LABELS = {
    MatchStatus.MATCHED: "Matched invoice and payment",
    MatchStatus.UNMATCHED_INVOICE: "Invoice missing payment",
    MatchStatus.UNMATCHED_PAYMENT: "Payment missing invoice",
    MatchStatus.AMOUNT_MISMATCH: "Payment amount differs",
    MatchStatus.CURRENCY_MISMATCH: "Payment currency differs",
    MatchStatus.AMBIGUOUS_REFERENCE: "Duplicate reference needs review",
}

AMBIGUOUS_REASON_LABELS = {
    AmbiguousReferenceReason.DUPLICATE_INVOICE_REFERENCE: (
        "Duplicate invoice reference; review invoice export duplicates"
    ),
    AmbiguousReferenceReason.DUPLICATE_PAYMENT_REFERENCE: (
        "Duplicate payment reference; review payment export duplicates"
    ),
    AmbiguousReferenceReason.DUPLICATE_INVOICE_AND_PAYMENT_REFERENCE: (
        "Duplicate invoice and payment reference; review both exports"
    ),
}

SUMMARY_COLUMNS = ("status", "status_label", "count")
DETAIL_COLUMNS = (
    "status",
    "status_label",
    "reference",
    "invoice_id",
    "payment_id",
    "customer_name",
    "invoice_amount",
    "payment_amount",
    "invoice_source_row",
    "payment_source_row",
    "reason",
)
WORKBOOK_DETAIL_COLUMNS = (
    "status",
    "status_label",
    "reference",
    "invoice_id",
    "payment_id",
    "customer_name",
    "invoice_amount",
    "invoice_currency",
    "payment_amount",
    "payment_currency",
    "invoice_source_row",
    "payment_source_row",
    "reason",
)
WORKBOOK_DETAIL_HEADERS = (
    "Status",
    "Status Label",
    "Reference",
    "Invoice ID",
    "Payment ID",
    "Customer",
    "Invoice Amount",
    "Invoice Currency",
    "Payment Amount",
    "Payment Currency",
    "Invoice Source Row",
    "Payment Source Row",
    "Reason",
)
WORKBOOK_SHEET_NAMES = (
    "Summary",
    "Matched",
    "Exceptions",
    "Invoice Exceptions",
    "Payment Exceptions",
    "Details",
)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
_HEADER_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=12)
_WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


@dataclass(frozen=True, slots=True)
class ReportOutputPaths:
    """Paths written by the local report writer."""

    markdown: Path
    summary_csv: Path
    details_csv: Path
    workbook_xlsx: Path


def build_summary_counts(result: ReconciliationResult) -> dict[MatchStatus, int]:
    """Return deterministic counts by reconciliation status."""
    counts_by_status = {
        MatchStatus.MATCHED: len(result.matched_pairs),
        MatchStatus.UNMATCHED_INVOICE: len(result.unmatched_invoices),
        MatchStatus.UNMATCHED_PAYMENT: len(result.unmatched_payments),
        MatchStatus.AMOUNT_MISMATCH: len(result.amount_mismatches),
        MatchStatus.CURRENCY_MISMATCH: len(result.currency_mismatches),
        MatchStatus.AMBIGUOUS_REFERENCE: len(result.ambiguous_references),
    }
    return {status: counts_by_status[status] for status in STATUS_ORDER}


def render_markdown_report(result: ReconciliationResult) -> str:
    """Render a deterministic client-readable Markdown report."""
    lines = [
        "# Invoice Payment Reconciliation Report",
        "",
        "## Reconciliation Totals",
        "",
    ]
    lines.extend(
        _markdown_table(
            ("Metric", "Value"),
            (
                (label, str(value))
                for label, value in _build_report_totals(result).items()
            ),
        )
    )
    lines.extend(["", "## Status Summary", ""])
    lines.extend(
        _markdown_table(
            ("Category", "Count"),
            [
                (STATUS_LABELS[status], str(count))
                for status, count in build_summary_counts(result).items()
            ],
        )
    )

    lines.extend(
        _markdown_section(
            "Matched Records",
            ("Reference", "Invoice ID", "Payment ID", "Customer", "Matched Amount"),
            (
                (
                    pair.reference,
                    pair.invoice.invoice_id,
                    pair.payment.payment_id,
                    pair.invoice.customer_name,
                    _format_money(pair.invoice.invoice_amount),
                )
                for pair in _sorted_matched_pairs(result.matched_pairs)
            ),
        )
    )
    lines.extend(
        _markdown_section(
            "Invoices Missing Payments",
            (
                "Reference",
                "Invoice ID",
                "Customer",
                "Invoice Amount",
                "Review Note",
            ),
            (
                (
                    item.reference,
                    item.invoice.invoice_id,
                    item.invoice.customer_name,
                    _format_money(item.invoice.invoice_amount),
                    _unmatched_invoice_note(),
                )
                for item in _sorted_unmatched_invoices(result.unmatched_invoices)
            ),
        )
    )
    lines.extend(
        _markdown_section(
            "Payments Missing Invoices",
            (
                "Reference",
                "Payment ID",
                "Customer",
                "Payment Amount",
                "Review Note",
            ),
            (
                (
                    item.reference,
                    item.payment.payment_id,
                    item.payment.customer_name,
                    _format_money(item.payment.payment_amount),
                    _unmatched_payment_note(),
                )
                for item in _sorted_unmatched_payments(result.unmatched_payments)
            ),
        )
    )
    lines.extend(
        _markdown_section(
            "Amount Variances (Underpaid/Overpaid)",
            (
                "Reference",
                "Invoice ID",
                "Payment ID",
                "Customer",
                "Invoice Amount",
                "Payment Amount",
                "Variance",
            ),
            (
                (
                    item.reference,
                    item.invoice.invoice_id,
                    item.payment.payment_id,
                    item.invoice.customer_name,
                    _format_money(item.invoice.invoice_amount),
                    _format_money(item.payment.payment_amount),
                    _amount_variance_note(item),
                )
                for item in _sorted_amount_mismatches(result.amount_mismatches)
            ),
        )
    )
    lines.extend(
        _markdown_section(
            "Currency Conflicts",
            (
                "Reference",
                "Invoice ID",
                "Payment ID",
                "Customer",
                "Invoice Amount",
                "Payment Amount",
                "Review Note",
            ),
            (
                (
                    item.reference,
                    item.invoice.invoice_id,
                    item.payment.payment_id,
                    item.invoice.customer_name,
                    _format_money(item.invoice.invoice_amount),
                    _format_money(item.payment.payment_amount),
                    _currency_mismatch_note(item),
                )
                for item in _sorted_currency_mismatches(result.currency_mismatches)
            ),
        )
    )
    lines.extend(
        _markdown_section(
            "Duplicate References Needing Review",
            (
                "Reference",
                "Invoice IDs",
                "Payment IDs",
                "Invoice Rows",
                "Payment Rows",
                "Review Note",
            ),
            (
                (
                    item.reference,
                    _join_values(invoice.invoice_id for invoice in item.invoices),
                    _join_values(payment.payment_id for payment in item.payments),
                    _join_values(str(invoice.source_row) for invoice in item.invoices),
                    _join_values(str(payment.source_row) for payment in item.payments),
                    AMBIGUOUS_REASON_LABELS[item.reason],
                )
                for item in _sorted_ambiguous_references(result.ambiguous_references)
            ),
        )
    )

    return "\n".join(lines).rstrip() + "\n"


def _build_report_totals(result: ReconciliationResult) -> dict[str, int]:
    status_counts = build_summary_counts(result)
    exception_groups = sum(
        count
        for status, count in status_counts.items()
        if status != MatchStatus.MATCHED
    )
    return {
        "Invoice records reviewed": _invoice_record_count(result),
        "Payment records reviewed": _payment_record_count(result),
        "Matched invoice/payment pairs": status_counts[MatchStatus.MATCHED],
        "Exception groups needing review": exception_groups,
    }


def _invoice_record_count(result: ReconciliationResult) -> int:
    return (
        len(result.matched_pairs)
        + len(result.unmatched_invoices)
        + len(result.amount_mismatches)
        + len(result.currency_mismatches)
        + sum(len(item.invoices) for item in result.ambiguous_references)
    )


def _payment_record_count(result: ReconciliationResult) -> int:
    return (
        len(result.matched_pairs)
        + len(result.unmatched_payments)
        + len(result.amount_mismatches)
        + len(result.currency_mismatches)
        + sum(len(item.payments) for item in result.ambiguous_references)
    )


def render_summary_csv(result: ReconciliationResult) -> str:
    """Render status counts as deterministic CSV text."""
    rows = [
        {
            "status": status.value,
            "status_label": STATUS_LABELS[status],
            "count": str(count),
        }
        for status, count in build_summary_counts(result).items()
    ]
    return _render_csv(SUMMARY_COLUMNS, rows)


def render_details_csv(result: ReconciliationResult) -> str:
    """Render reconciliation detail rows as deterministic CSV text."""
    return _render_csv(DETAIL_COLUMNS, _detail_rows(result))


def write_reconciliation_reports(
    result: ReconciliationResult,
    output_dir: str | Path,
) -> ReportOutputPaths:
    """Write Markdown, CSV, and XLSX reconciliation reports to a local directory."""
    directory = Path(output_dir)
    directory.mkdir(parents=True, exist_ok=True)

    paths = ReportOutputPaths(
        markdown=directory / "reconciliation-report.md",
        summary_csv=directory / "reconciliation-summary.csv",
        details_csv=directory / "reconciliation-details.csv",
        workbook_xlsx=directory / "reconciliation-workbook.xlsx",
    )
    _write_text(paths.markdown, render_markdown_report(result))
    _write_text(paths.summary_csv, render_summary_csv(result))
    _write_text(paths.details_csv, render_details_csv(result))
    write_reconciliation_workbook(result, paths.workbook_xlsx)
    return paths


def write_reconciliation_workbook(
    result: ReconciliationResult,
    path: str | Path,
) -> Path:
    """Write a deterministic XLSX workbook for reviewer-friendly reconciliation."""
    output_path = Path(path)
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = WORKBOOK_SHEET_NAMES[0]
    _write_summary_sheet(summary_sheet, result)

    workbook_rows = _workbook_rows(result)
    _write_table_sheet(
        workbook.create_sheet(WORKBOOK_SHEET_NAMES[1]),
        WORKBOOK_DETAIL_HEADERS,
        workbook_rows["matched"],
    )
    _write_table_sheet(
        workbook.create_sheet(WORKBOOK_SHEET_NAMES[2]),
        WORKBOOK_DETAIL_HEADERS,
        workbook_rows["exceptions"],
    )
    _write_table_sheet(
        workbook.create_sheet(WORKBOOK_SHEET_NAMES[3]),
        WORKBOOK_DETAIL_HEADERS,
        workbook_rows["invoice_exceptions"],
    )
    _write_table_sheet(
        workbook.create_sheet(WORKBOOK_SHEET_NAMES[4]),
        WORKBOOK_DETAIL_HEADERS,
        workbook_rows["payment_exceptions"],
    )
    _write_table_sheet(
        workbook.create_sheet(WORKBOOK_SHEET_NAMES[5]),
        WORKBOOK_DETAIL_HEADERS,
        workbook_rows["details"],
    )

    workbook.save(output_path)
    return output_path


def _detail_rows(result: ReconciliationResult) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    rows.extend(
        _matched_row(pair) for pair in _sorted_matched_pairs(result.matched_pairs)
    )
    rows.extend(
        _unmatched_invoice_row(item)
        for item in _sorted_unmatched_invoices(result.unmatched_invoices)
    )
    rows.extend(
        _unmatched_payment_row(item)
        for item in _sorted_unmatched_payments(result.unmatched_payments)
    )
    rows.extend(
        _amount_mismatch_row(item)
        for item in _sorted_amount_mismatches(result.amount_mismatches)
    )
    rows.extend(
        _currency_mismatch_row(item)
        for item in _sorted_currency_mismatches(result.currency_mismatches)
    )
    rows.extend(
        _ambiguous_reference_row(item)
        for item in _sorted_ambiguous_references(result.ambiguous_references)
    )
    return rows


def _workbook_rows(result: ReconciliationResult) -> dict[str, list[tuple[str, ...]]]:
    matched_rows = [
        _workbook_matched_row(pair)
        for pair in _sorted_matched_pairs(result.matched_pairs)
    ]
    unmatched_invoice_rows = [
        _workbook_unmatched_invoice_row(item)
        for item in _sorted_unmatched_invoices(result.unmatched_invoices)
    ]
    unmatched_payment_rows = [
        _workbook_unmatched_payment_row(item)
        for item in _sorted_unmatched_payments(result.unmatched_payments)
    ]
    amount_mismatch_rows = [
        _workbook_amount_mismatch_row(item)
        for item in _sorted_amount_mismatches(result.amount_mismatches)
    ]
    currency_mismatch_rows = [
        _workbook_currency_mismatch_row(item)
        for item in _sorted_currency_mismatches(result.currency_mismatches)
    ]
    ambiguous_references = _sorted_ambiguous_references(result.ambiguous_references)
    ambiguous_rows = [
        _workbook_ambiguous_reference_row(item) for item in ambiguous_references
    ]
    exception_rows = (
        unmatched_invoice_rows
        + unmatched_payment_rows
        + amount_mismatch_rows
        + currency_mismatch_rows
        + ambiguous_rows
    )
    invoice_exception_rows = (
        unmatched_invoice_rows
        + amount_mismatch_rows
        + currency_mismatch_rows
        + [
            _workbook_ambiguous_reference_row(item)
            for item in ambiguous_references
            if item.invoices
        ]
    )
    payment_exception_rows = (
        unmatched_payment_rows
        + amount_mismatch_rows
        + currency_mismatch_rows
        + [
            _workbook_ambiguous_reference_row(item)
            for item in ambiguous_references
            if item.payments
        ]
    )

    return {
        "matched": matched_rows,
        "exceptions": exception_rows,
        "invoice_exceptions": invoice_exception_rows,
        "payment_exceptions": payment_exception_rows,
        "details": matched_rows + exception_rows,
    }


def _matched_row(item: MatchedPair) -> dict[str, str]:
    row = _base_detail_row(item.status, item.reference)
    _add_invoice(row, item.invoice)
    _add_payment(row, item.payment)
    return row


def _unmatched_invoice_row(item: UnmatchedInvoice) -> dict[str, str]:
    row = _base_detail_row(item.status, item.reference)
    _add_invoice(row, item.invoice)
    row["reason"] = _unmatched_invoice_note()
    return row


def _unmatched_payment_row(item: UnmatchedPayment) -> dict[str, str]:
    row = _base_detail_row(item.status, item.reference)
    _add_payment(row, item.payment)
    row["reason"] = _unmatched_payment_note()
    return row


def _amount_mismatch_row(item: AmountMismatch) -> dict[str, str]:
    row = _base_detail_row(item.status, item.reference)
    _add_invoice(row, item.invoice)
    _add_payment(row, item.payment)
    row["reason"] = _amount_variance_note(item)
    return row


def _currency_mismatch_row(item: CurrencyMismatch) -> dict[str, str]:
    row = _base_detail_row(item.status, item.reference)
    _add_invoice(row, item.invoice)
    _add_payment(row, item.payment)
    row["reason"] = _currency_mismatch_note(item)
    return row


def _ambiguous_reference_row(item: AmbiguousReference) -> dict[str, str]:
    row = _base_detail_row(item.status, item.reference)
    row["invoice_id"] = _join_values(invoice.invoice_id for invoice in item.invoices)
    row["payment_id"] = _join_values(payment.payment_id for payment in item.payments)
    row["customer_name"] = _join_values(
        record.customer_name for record in (*item.invoices, *item.payments)
    )
    row["invoice_amount"] = _join_values(
        _format_money(invoice.invoice_amount) for invoice in item.invoices
    )
    row["payment_amount"] = _join_values(
        _format_money(payment.payment_amount) for payment in item.payments
    )
    row["invoice_source_row"] = _join_values(
        str(invoice.source_row) for invoice in item.invoices
    )
    row["payment_source_row"] = _join_values(
        str(payment.source_row) for payment in item.payments
    )
    row["reason"] = AMBIGUOUS_REASON_LABELS[item.reason]
    return row


def _workbook_matched_row(item: MatchedPair) -> tuple[str, ...]:
    return _workbook_detail_tuple(_matched_row(item))


def _workbook_unmatched_invoice_row(item: UnmatchedInvoice) -> tuple[str, ...]:
    return _workbook_detail_tuple(_unmatched_invoice_row(item))


def _workbook_unmatched_payment_row(item: UnmatchedPayment) -> tuple[str, ...]:
    return _workbook_detail_tuple(_unmatched_payment_row(item))


def _workbook_amount_mismatch_row(item: AmountMismatch) -> tuple[str, ...]:
    return _workbook_detail_tuple(_amount_mismatch_row(item))


def _workbook_currency_mismatch_row(item: CurrencyMismatch) -> tuple[str, ...]:
    return _workbook_detail_tuple(_currency_mismatch_row(item))


def _workbook_ambiguous_reference_row(item: AmbiguousReference) -> tuple[str, ...]:
    return _workbook_detail_tuple(_ambiguous_reference_row(item))


def _workbook_detail_tuple(row: dict[str, str]) -> tuple[str, ...]:
    invoice_amount, invoice_currency = _split_money_text(row["invoice_amount"])
    payment_amount, payment_currency = _split_money_text(row["payment_amount"])
    enriched_row = {
        **row,
        "invoice_amount": invoice_amount,
        "invoice_currency": invoice_currency,
        "payment_amount": payment_amount,
        "payment_currency": payment_currency,
    }
    return tuple(enriched_row[column] for column in WORKBOOK_DETAIL_COLUMNS)


def _base_detail_row(status: MatchStatus, reference: str) -> dict[str, str]:
    return {
        "status": status.value,
        "status_label": STATUS_LABELS[status],
        "reference": reference,
        "invoice_id": "",
        "payment_id": "",
        "customer_name": "",
        "invoice_amount": "",
        "payment_amount": "",
        "invoice_source_row": "",
        "payment_source_row": "",
        "reason": "",
    }


def _add_invoice(row: dict[str, str], invoice: InvoiceRecord) -> None:
    row["invoice_id"] = invoice.invoice_id
    row["customer_name"] = invoice.customer_name
    row["invoice_amount"] = _format_money(invoice.invoice_amount)
    row["invoice_source_row"] = str(invoice.source_row)


def _add_payment(row: dict[str, str], payment: PaymentRecord) -> None:
    row["payment_id"] = payment.payment_id
    row["customer_name"] = row["customer_name"] or payment.customer_name
    row["payment_amount"] = _format_money(payment.payment_amount)
    row["payment_source_row"] = str(payment.source_row)


def _unmatched_invoice_note() -> str:
    return "No payment found for invoice reference"


def _unmatched_payment_note() -> str:
    return "No invoice found for payment reference"


def _amount_variance_note(item: AmountMismatch) -> str:
    invoice_amount = item.invoice.invoice_amount
    payment_amount = item.payment.payment_amount
    variance = payment_amount.value - invoice_amount.value
    formatted_variance = _format_money(
        MoneyAmount(value=abs(variance), currency=invoice_amount.currency)
    )
    if variance < 0:
        return f"Underpaid by {formatted_variance}; possible partial payment"
    return f"Overpaid by {formatted_variance}"


def _currency_mismatch_note(item: CurrencyMismatch) -> str:
    invoice_currency = item.invoice.invoice_amount.currency
    payment_currency = item.payment.payment_amount.currency
    return f"Invoice currency {invoice_currency}; payment currency {payment_currency}"


def _sorted_matched_pairs(items: Iterable[MatchedPair]) -> list[MatchedPair]:
    return sorted(
        items,
        key=lambda item: (
            _text_key(item.reference),
            _text_key(item.invoice.invoice_id),
            _text_key(item.payment.payment_id),
            item.invoice.source_row,
            item.payment.source_row,
        ),
    )


def _sorted_unmatched_invoices(
    items: Iterable[UnmatchedInvoice],
) -> list[UnmatchedInvoice]:
    return sorted(
        items,
        key=lambda item: (
            _text_key(item.reference),
            _text_key(item.invoice.invoice_id),
            item.invoice.source_row,
        ),
    )


def _sorted_unmatched_payments(
    items: Iterable[UnmatchedPayment],
) -> list[UnmatchedPayment]:
    return sorted(
        items,
        key=lambda item: (
            _text_key(item.reference),
            _text_key(item.payment.payment_id),
            item.payment.source_row,
        ),
    )


def _sorted_amount_mismatches(
    items: Iterable[AmountMismatch],
) -> list[AmountMismatch]:
    return sorted(
        items,
        key=lambda item: (
            _text_key(item.reference),
            _text_key(item.invoice.invoice_id),
            _text_key(item.payment.payment_id),
            item.invoice.source_row,
            item.payment.source_row,
        ),
    )


def _sorted_currency_mismatches(
    items: Iterable[CurrencyMismatch],
) -> list[CurrencyMismatch]:
    return sorted(
        items,
        key=lambda item: (
            _text_key(item.reference),
            _text_key(item.invoice.invoice_id),
            _text_key(item.payment.payment_id),
            item.invoice.source_row,
            item.payment.source_row,
        ),
    )


def _sorted_ambiguous_references(
    items: Iterable[AmbiguousReference],
) -> list[AmbiguousReference]:
    return sorted(
        items,
        key=lambda item: (
            _text_key(item.reference),
            tuple(
                (_text_key(invoice.invoice_id), invoice.source_row)
                for invoice in item.invoices
            ),
            tuple(
                (_text_key(payment.payment_id), payment.source_row)
                for payment in item.payments
            ),
        ),
    )


def _text_key(value: str) -> tuple[str, str]:
    return (value.casefold(), value)


def _markdown_section(
    title: str,
    headers: tuple[str, ...],
    rows: Iterable[tuple[str, ...]],
) -> list[str]:
    row_list = list(rows)
    if not row_list:
        return []
    section = ["", f"## {title}", ""]
    section.extend(_markdown_table(headers, row_list))
    return section


def _markdown_table(
    headers: tuple[str, ...],
    rows: Iterable[tuple[str, ...]],
) -> list[str]:
    table = [
        "| " + " | ".join(_markdown_cell(header) for header in headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    table.extend(
        "| " + " | ".join(_markdown_cell(value) for value in row) + " |" for row in rows
    )
    return table


def _markdown_cell(value: str) -> str:
    return value.replace("|", "\\|")


def _render_csv(columns: tuple[str, ...], rows: Iterable[dict[str, str]]) -> str:
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=columns, lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue()


def _write_summary_sheet(worksheet, result: ReconciliationResult) -> None:
    worksheet.append(("Metric", "Value"))
    _style_header_row(worksheet, 1)
    for metric, value in _build_report_totals(result).items():
        worksheet.append((metric, value))

    worksheet.append(("", ""))
    status_header_row = worksheet.max_row + 1
    worksheet.append(("Status", "Status Label", "Count"))
    _style_header_row(worksheet, status_header_row)
    for status, count in build_summary_counts(result).items():
        worksheet.append((status.value, STATUS_LABELS[status], count))

    worksheet.freeze_panes = "A2"
    _apply_column_widths(worksheet)


def _write_table_sheet(
    worksheet,
    headers: tuple[str, ...],
    rows: Iterable[tuple[str, ...]],
) -> None:
    worksheet.append(headers)
    _style_header_row(worksheet, 1)
    for row in rows:
        worksheet.append(row)

    worksheet.freeze_panes = "A2"
    if worksheet.max_row > 1:
        last_column = get_column_letter(worksheet.max_column)
        worksheet.auto_filter.ref = f"A1:{last_column}{worksheet.max_row}"
    _apply_column_widths(worksheet)


def _style_header_row(worksheet, row_number: int) -> None:
    for cell in worksheet[row_number]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP_ALIGNMENT


def _apply_column_widths(worksheet) -> None:
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            cell.alignment = _WRAP_ALIGNMENT
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(
                max_length,
                max((len(part) for part in cell_value.splitlines()), default=0),
            )
        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            38,
        )


def _write_text(path: Path, content: str) -> None:
    with path.open("w", encoding="utf-8", newline="\n") as file:
        file.write(content)


def _format_money(money: MoneyAmount) -> str:
    return f"{_format_decimal(money.value)} {money.currency}"


def _format_decimal(value: Decimal) -> str:
    return format(value, "f")


def _join_values(values: Iterable[str]) -> str:
    unique_values: list[str] = []
    for value in values:
        if value and value not in unique_values:
            unique_values.append(value)
    return "; ".join(unique_values)


def _split_money_text(value: str) -> tuple[str, str]:
    if not value:
        return "", ""

    amounts: list[str] = []
    currencies: list[str] = []
    for item in value.split("; "):
        if " " not in item:
            amounts.append(item)
            continue
        amount, currency = item.rsplit(" ", 1)
        amounts.append(amount)
        if currency not in currencies:
            currencies.append(currency)

    return "; ".join(amounts), "; ".join(currencies)
